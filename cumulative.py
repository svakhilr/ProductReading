from openpyxl import load_workbook
from collections import defaultdict


workbook = load_workbook("product_view_analysis.xlsx")
sheet = workbook.active


headers = [cell.value.strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
print(headers)

parent_org_idx = headers.index("parent org")
brand_idx = headers.index("brand")
view_count_idx = headers.index("product view count")


data = defaultdict(lambda: defaultdict(int))


for row in sheet.iter_rows(min_row=2, values_only=True):
    parent_org = str(row[parent_org_idx]).strip()
    brand = str(row[brand_idx]).strip()
    try:
        view_count = int(float(row[view_count_idx]))
    except (ValueError, TypeError):
        view_count = 0
    data[parent_org][brand] += view_count


for parent_org, brands in data.items():
    print(f"{parent_org}:")
    sorted_brands = sorted(brands.items(), key=lambda x: x[1], reverse=False)
    for brand, count in sorted_brands:
        print(f"  {brand}: {count}")

